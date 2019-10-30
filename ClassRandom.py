#用于出现麻烦事情的时候随机到倒霉蛋：
#@author skelviper

from openpyxl import load_workbook
import random
from random import choice

def readExcel(location):    #读取文件sheet1的第一列，存入data列表
    book = load_workbook(filename=location)
    #sheet = book.get_sheet_by_name("Sheet1") o
    #this is too old,cause warning
    sheet = book.worksheets[0]
    data= []
    row_num = 1
    while sheet.cell(row=row_num, column=1).value :
        data.append(sheet.cell(row=row_num, column=1).value)
        row_num = row_num + 1
    return data

location = r'namelist.xlsx'
data =  readExcel(location)

num =8

print(random.sample(data,num))